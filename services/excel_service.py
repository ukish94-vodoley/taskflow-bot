from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORTS_DIR = Path(__file__).resolve().parent.parent / "reports"


def create_finance_report(rows, period_label: str) -> Path:
    REPORTS_DIR.mkdir(exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Moliya harakati"

    sheet.merge_cells("A1:G1")
    sheet["A1"] = "TaskFlow — moliyaviy hisobot"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="162B4D")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A2:G2")
    sheet["A2"] = f"Davr: {period_label} | Tayyorlangan vaqt: {datetime.now():%d.%m.%Y %H:%M}"
    sheet["A2"].font = Font(italic=True, color="667085")

    headers = ["№", "Sana", "Xodim", "Rahbar", "Harakat", "Summa (so‘m)", "Izoh"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="246BFE")
        cell.alignment = Alignment(horizontal="center")

    for index, (finance, employee_name, leader_name) in enumerate(rows, start=1):
        row = index + 4
        sheet.cell(row=row, column=1, value=index)
        sheet.cell(row=row, column=2, value=finance.created_at)
        sheet.cell(row=row, column=2).number_format = "dd.mm.yyyy hh:mm"
        sheet.cell(row=row, column=3, value=employee_name)
        sheet.cell(row=row, column=4, value=leader_name)
        sheet.cell(row=row, column=5, value="To‘ldirish" if finance.type == "topup" else "Xarajat")
        amount_cell = sheet.cell(row=row, column=6, value=finance.amount if finance.type == "topup" else -finance.amount)
        amount_cell.number_format = '#,##0'
        amount_cell.font = Font(color="087B54" if finance.type == "topup" else "B33333")
        sheet.cell(row=row, column=7, value=finance.description or "")

    summary = workbook.create_sheet("Xulosa")
    summary["A1"] = "TaskFlow — xulosa"
    summary["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="162B4D")
    summary.merge_cells("A1:C1")
    summary["A3"] = "Davr"
    summary["B3"] = period_label
    summary["A4"] = "Jami to‘ldirish"
    summary["B4"] = sum(finance.amount for finance, _, _ in rows if finance.type == "topup")
    summary["A5"] = "Jami xarajat"
    summary["B5"] = sum(finance.amount for finance, _, _ in rows if finance.type == "expense")
    summary["A6"] = "Sof qoldiq"
    summary["B6"] = summary["B4"].value - summary["B5"].value
    for cell in (summary["B4"], summary["B5"], summary["B6"]):
        cell.number_format = '#,##0'
        cell.font = Font(bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A5" if worksheet.title == "Moliya harakati" else "A3"
        for column_cells in worksheet.columns:
            index = column_cells[0].column
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 14), 42)

    filename = f"taskflow-moliya-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    output = REPORTS_DIR / filename
    workbook.save(output)
    return output
